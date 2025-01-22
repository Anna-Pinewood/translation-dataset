from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import logging
from typing import List, Dict, Optional
from models import TranslationEntry
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import Border, Side


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Handles Excel export/import operations for translation data"""

    REQUIRED_COLUMNS = [
        'idx', 'prompt', 'prompt_ru',
        'response', 'response_ru', 'is_response_unsecure'
    ]

    def __init__(self):
        """Initialize ExcelExporter"""
        logger.info("Initializing ExcelExporter")

    def export_to_excel(
        self,
        entries: List[TranslationEntry],
        output_path: Path,
        translated_only: bool = True
    ) -> None:
        """
        Export translation entries to Excel file.

        Parameters
        ----------
        entries : List[TranslationEntry]
            List of entries to export
        output_path : Path
            Path to save Excel file
        translated_only : bool, optional
            If True, export only translated entries, by default True
        """
        logger.info("Exporting entries to Excel: %s", output_path)

        # Convert entries to DataFrame format
        data = []
        for entry in entries:
            entry_dict = entry.to_dict()
            # Add is_response_unsecure field if not present
            if 'is_response_unsecure' not in entry_dict:
                entry_dict['is_response_unsecure'] = False

            # Filter out untranslated entries if requested
            if translated_only:
                if not entry_dict['prompt_ru'] or not entry_dict['response_ru']:
                    continue

            data.append(entry_dict)

        if not data:
            logger.warning("No entries to export")
            return

        df = pd.DataFrame(data)

        # Ensure all required columns are present
        for col in self.REQUIRED_COLUMNS:
            if col not in df.columns:
                df[col] = None

        # Reorder columns to match required order
        df = df[self.REQUIRED_COLUMNS]

        # Export to Excel
        df.to_excel(output_path, index=False)
        self.prettify_excel(output_path)
        logger.info("Successfully exported %d entries to Excel", len(df))

    def prettify_excel(self, excel_path: str):
        wb = load_workbook(excel_path)
        ws = wb.active  # Get the active sheet
        column_widths = {
            "A": 10,  # idx
            "B": 60,  # prompt
            "C": 60,  # prompt_ru
            "D": 60,  # response
            "E": 60   # response_ru
        }
        bg_fill = PatternFill(start_color="C4E3EF",
                              end_color="C4E3EF", fill_type="solid")
        highlight_columns = ["C", "E"]  # A -> prompt_ru, C -> response_ru

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width  # Set column width

        for row in ws.iter_rows():
            for cell in row:
                # Apply text alignment
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                # Set font size for all cells
                cell.font = Font(size=14)
                # Set bg color
                if cell.column_letter in highlight_columns:
                    cell.fill = bg_fill

        for cell in ws[1]:  # Iterate through the first row (header row)
            # Make the font bold and set size
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(
                horizontal="center", vertical="center")  # Center the text

        thin_border = Border(left=Side(style="thin"), right=Side(
            style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Columns A-F (1-6)
        for row in ws.iter_rows(min_col=1, max_col=6, min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

        logger.info("Prettified Excel file")
        wb.save(excel_path)

    def update_from_excel(
        self,
        excel_path: Path,
        current_entries: Dict[int, TranslationEntry]
    ) -> List[TranslationEntry]:
        """
        Update entries from Excel file.

        Parameters
        ----------
        excel_path : Path
            Path to Excel file with updated translations
        current_entries : Dict[int, TranslationEntry]
            Dictionary of current entries mapped by their indices

        Returns
        -------
        List[TranslationEntry]
            List of updated entries

        Raises
        ------
        ValueError
            If Excel file structure is invalid or entries don't match
        """
        logger.info("Updating entries from Excel: %s", excel_path)

        # Read Excel file
        df = pd.read_excel(excel_path)

        # Validate structure
        missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
        if missing_cols:
            raise ValueError(
                f"Excel file missing required columns: {missing_cols}")

        # Process updates
        updated_entries = []
        for _, row in df.iterrows():
            idx = row['idx']

            # Verify entry exists
            if idx not in current_entries:
                logger.warning(
                    f"Entry with idx {idx} not found in current entries")
                continue

            # Get current entry
            entry = current_entries[idx]

            # Update translations and security flag
            entry.set_translation(
                prompt_ru=row['prompt_ru'],
                response_ru=row['response_ru']
            )

            # Update security flag if changed
            if 'is_response_unsecure' in row:
                if row['is_response_unsecure'] != getattr(entry, 'is_response_unsecure', False):
                    setattr(entry, 'is_response_unsecure',
                            bool(row['is_response_unsecure']))

            updated_entries.append(entry)

        logger.info("Successfully processed %d entries from Excel",
                    len(updated_entries))
        return updated_entries
